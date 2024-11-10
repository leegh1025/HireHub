from django.db.models import Q, Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from applicants.models import Application, Answer
from django.db import models, transaction
from .tasks import process_application
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import datetime
from django.contrib.auth.decorators import login_required

import io
import zipfile
import xlwt
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, Http404
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch


from django.core.mail import send_mail
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth import logout as auth_logout

# 인증 코드 생성
import json
from django.utils.crypto import get_random_string
from .models import VerificationCode

# 비밀번호 재설정
from django.contrib.auth.views import PasswordResetView
from django.views import View
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.urls import reverse

from .models import Applicant, Application, Answer, Possible_date_list, Comment, individualQuestion, individualAnswer, Interviewer, AudioRecording
from accounts.models import Interviewer, InterviewTeam
from template.models import ApplicationTemplate, ApplicationQuestion, InterviewTemplate, InterviewQuestion, EvaluationTemplate
from .forms import ApplicationForm, CommentForm, QuestionForm, AnswerForm, ApplyForm

from template.models import ApplicationTemplate, InterviewTemplate, InterviewQuestion
from .forms import CommentForm, QuestionForm, AnswerForm, ApplyForm, CustomPasswordResetForm

# 지원자 초기 페이지
def initial(request):
    template = ApplicationTemplate.objects.get(is_default='1') # pk 변경 필요
    # 목표 시간을 설정합니다.
    target_time = timezone.make_aware(datetime(2024, 12, 1, 16, 0, 0), timezone=timezone.get_current_timezone())
    print(target_time)
    # 현재 시간 가져오기
    current_time = timezone.localtime(timezone.now())
    print(current_time)
    # 목표 시간을 지났는지 여부를 계산
    time_over = current_time >= target_time
    print(time_over)

    context = {'template': template, 'time_over': time_over,}
    return render(request, 'for_applicant/initial.html', context)

# 지원자 회원가입
def signup(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        email = data.get('email')
        name = data.get('name')
        phone_number = data.get('phone_number')
        password = data.get('password')
        password_confirm = data.get('password_confirm')
        verification_code = data.get('code')

        if Applicant.objects.filter(phone_number=phone_number).exists():
            return JsonResponse({'success': False, 'message': '이미 가입된 전화번호입니다.'})

        if password != password_confirm:
            return JsonResponse({'success': False, 'message': '비밀번호가 일치하지 않습니다.'})

        try:
            verification = VerificationCode.objects.get(email=email)

            if not verification.is_verified:
                return JsonResponse({'success': False, 'message': '이메일 인증을 완료해야 합니다.'})

            if verification.code == verification_code and not verification.is_expired():
                applicant = Applicant.objects.create_user(
                    email=email,
                    name=name,
                    phone_number=phone_number,
                    password=password
                )
                return JsonResponse({'success': True, 'redirect_url': '/applicants/'})
            
            return JsonResponse({'success': False, 'message': '유효하지 않은 인증번호이거나 인증번호가 만료되었습니다.'})

        except VerificationCode.DoesNotExist:
            return JsonResponse({'success': False, 'message': '입력한 이메일에 대한 인증번호가 존재하지 않습니다.'})
    
    return render(request, 'for_applicant/signup.html')
    
def send_verification_code(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        
        if not email:
            return JsonResponse({'success': False, 'message': '이메일을 입력해주세요.'})
        if Applicant.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': '이미 존재하는 이메일입니다.'})

        # 인증번호 생성 및 이메일 발송
        code = get_random_string(length=6, allowed_chars='0123456789')
        verification, created = VerificationCode.objects.update_or_create(
            email=email,
            defaults={'code': code, 'created_at': timezone.now()}  # 새로운 인증번호와 생성 시간 업데이트
        )

        send_mail(
            'Your verification code',
            f'Your verification code is {code}',
            'pirogramming@naver.com',
            [email],
            fail_silently=False,
        )

        return JsonResponse({'success': True, 'message': '인증번호가 이메일로 전송되었습니다.'})

def verify_code(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        input_code = data.get('code')

        try:
            verification = VerificationCode.objects.get(email=email)
            
            verification_code = str(verification.code)
            input_code = str(input_code)
            if verification_code == input_code and not verification.is_expired():
                verification.is_verified = True
                verification.save()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'message': '유효하지 않은 인증번호이거나 인증번호가 만료되었습니다.'})
        
        except VerificationCode.DoesNotExist:
            return JsonResponse({'success': False, 'message': '인증번호가 입력되지 않았습니다.'})

def login(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            return JsonResponse({'success': False, 'message': '이메일과 비밀번호를 입력하세요.'})

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if user.is_active:
                auth_login(request, user)
                if isinstance(user, Applicant):
                    return JsonResponse({'success': True, 'redirect_url': '/applicants/'})
                else:
                    return JsonResponse({'success': False, 'message': '잘못된 계정입니다.'})
            else:
                return JsonResponse({'success': False, 'message': '계정이 비활성화되었습니다.'})
        else:
            return JsonResponse({'success': False, 'message': '이메일 또는 비밀번호가 틀렸습니다.'})

    return render(request, 'for_applicant/login.html')

def logout(request):
    auth_logout(request)
    return redirect('applicants:initial')

# 비밀번호 재설정
class ApplicantPasswordResetView(PasswordResetView):
    template_name = 'for_applicant/password_reset.html'
    email_template_name = 'for_applicant/password_reset_email.html'
    form_class = CustomPasswordResetForm

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            email = data.get('email', None)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'})
        
        # 폼에 데이터를 바인딩해서 유효성 검사
        form = self.form_class(data={'email': email})
        if form.is_valid():
            applicant = Applicant.objects.get(email=form.cleaned_data['email'])

            # uid와 token 생성
            uid = urlsafe_base64_encode(force_bytes(applicant.pk))
            token = default_token_generator.make_token(applicant)

            # 비밀번호 재설정 URL 생성
            reset_url = reverse('applicants:password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
            full_reset_url = f"{request.scheme}://{request.get_host()}{reset_url}"

            message = """
                비밀번호를 재설정하려면 아래 링크를 클릭하세요:
                <a href="{0}">비밀번호 재설정 링크</a>
            """.format(full_reset_url)

            send_mail(
                subject='비밀번호 재설정 요청',
                message='비밀번호 재설정 링크: {0}'.format(full_reset_url),  # 텍스트 버전
                from_email='pirogramming@naver.com',
                recipient_list=[email],
                fail_silently=False,
                html_message=message  # HTML 버전
            )

            return JsonResponse({'success': True, 'message': '비밀번호 재설정 이메일이 발송되었습니다.'})
        else:
            return JsonResponse({'success': False, 'message': '해당 이메일을 찾을 수 없습니다.'})

class ApplicantPasswordResetConfirmView(View):
    template_name = 'for_applicant/password_reset_confirm.html'
    
    def get(self, request, uidb64, token, *args, **kwargs):
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = Applicant.objects.get(pk=uid)
            if default_token_generator.check_token(user, token):
                return render(request, self.template_name)
            else:
                return render(request, 'for_applicant/password_reset_invalid.html')
        except (TypeError, ValueError, OverflowError, Applicant.DoesNotExist):
            return render(request, 'for_applicant/password_reset_invalid.html')
    
    def post(self, request, uidb64, token, *args, **kwargs):
        try:
            # JSON 데이터 파싱
            data = json.loads(request.body)
            new_password1 = data.get('new_password1')
            new_password2 = data.get('new_password2')

            # 기본 유효성 검사
            if not new_password1 or not new_password2:
                return JsonResponse({'success': False, 'message': '모든 필드를 입력해주세요.'})

            if new_password1 != new_password2:
                return JsonResponse({'success': False, 'message': '비밀번호가 일치하지 않습니다.'})

            # 사용자 확인
            try:
                uid = force_str(urlsafe_base64_decode(uidb64))
                user = Applicant.objects.get(pk=uid)
            except (TypeError, ValueError, OverflowError, Applicant.DoesNotExist):
                return JsonResponse({'success': False, 'message': '유효하지 않은 사용자입니다.'})

            # 토큰 확인
            if default_token_generator.check_token(user, token):
                user.set_password(new_password1)
                user.save()
                return JsonResponse({'success': True, 'message': '비밀번호가 성공적으로 변경되었습니다!', 'redirect_url': '/applicants/'})
            else:
                return JsonResponse({'success': False, 'message': '비밀번호 재설정 링크가 만료되었습니다.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': '잘못된 요청 형식입니다.'})  
        

def interview(request):
    if request.user.is_authenticated:
        applicants = Application.objects.all()
        ctx = {"applicants": applicants}
        return render(request, "applicant/interview.html", ctx)
    else:
        return redirect("accounts:login")

def change_status(request, status_zone_id, applicant_id):
    if request.method == 'POST':
        try:
            applicant = Application.objects.get(id=applicant_id)
            if status_zone_id == '1':
                applicant.status = 'interview_scheduled'
            elif status_zone_id == '2':
                applicant.status = 'interview_in_progress'
            elif status_zone_id == '3':
                applicant.status = 'interview_completed'
            applicant.save()
            return JsonResponse({'message': 'Status updated successfully'})
        except Applicant.DoesNotExist:
            return JsonResponse({'error': 'Applicant not found'}, status=404)
        except StatusZone.DoesNotExist:
            return JsonResponse({'error': 'Status zone not found'}, status=404)
    return JsonResponse({'error': 'Invalid request method'}, status=400)
            
def document(request):
    if request.user.is_authenticated:
        applicants = Application.objects.all()
        ctx = {"applicants": applicants}
        return render(request, "applicant/document.html", ctx)
    else:
        return redirect("accounts:login")

def search_applicant(request):
    search_txt = request.GET.get('search_txt')
    applicants = Application.objects.filter(name__icontains=search_txt)
    applicants = applicants.filter(~Q(status = 'submitted'))
    results = [{'id': applicant.id, 'name': applicant.name} for applicant in applicants]
    return JsonResponse(results, safe=False)



def custom_pagination(text, canvas_obj, x_position, y_position, max_width, line_height, font_name, font_size, bottom_margin):
    """
    텍스트를 페이지 너비에 맞게 줄바꿈하고, 페이지가 끝나면 자동으로 페이지를 넘겨가며 텍스트를 그립니다.
    """

    canvas_obj.setFont(font_name, font_size)
    words = text.split(' ')
    line = ''
    page_width, page_height = letter
    
    for word in words:
        test_line = f"{line} {word}".strip()
        width = canvas_obj.stringWidth(test_line, font_name, font_size)

        if width <= max_width:
            line = test_line
        else:
            canvas_obj.drawString(x_position, y_position, line)
            y_position -= line_height

            if y_position < bottom_margin:
                canvas_obj.showPage()  
                canvas_obj.setFont(font_name, font_size)  
                y_position = page_height - inch  

            line = word

    if line:
        canvas_obj.drawString(x_position, y_position, line)
        y_position -= line_height

    return y_position


def generate_pdf(applicant):
    pdf_buffer = io.BytesIO()
    p = canvas.Canvas(pdf_buffer, pagesize=letter)

    # 한글 폰트 등록
    pdfmetrics.registerFont(TTFont('Pretendard', 'static/fonts/PretendardVariable.ttf'))
    font_name = 'Pretendard'
    font_size = 10
    p.setFont(font_name, font_size)

    page_width, page_height = letter  
    margin = 40  
    max_width = page_width - 2 * margin  
    line_height = 18  
    bottom_margin = 50  

    # 지원자 기본 정보 작성
    y_position = 760  
    y_position = custom_pagination(
        f"지원자 정보| {str(applicant.name)} {str(applicant.school)}({str(applicant.major)})", 
        p, margin, y_position, max_width, line_height, font_name, font_size, bottom_margin
    )
    submission_date = applicant.submission_date.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
    y_position = custom_pagination(
        f"전화번호: {str(applicant.phone_number)} | 제출 일자: {submission_date}", 
        p, margin, y_position, max_width, line_height, font_name, font_size, bottom_margin
    )
    y_position -= 40  

    # 질문, 답변 작성
    for answer in applicant.answers.all():  
        question_text = f"[문항 {answer.question.id}] {answer.question.question_text}" 
        answer_text = str(answer.answer_text) 
        y_position = custom_pagination(
            question_text, p, margin, y_position, max_width, line_height, font_name, font_size, bottom_margin
        )
        y_position -= 10 
        y_position = custom_pagination(
            answer_text, p, margin, y_position, max_width, line_height, font_name, font_size, bottom_margin
        )
        y_position -= 20  

    p.showPage()
    p.save()
    pdf_buffer.seek(0)

    return pdf_buffer.getvalue()


# 1명 지원자 PDF 다운로드
def download_pdf_single(request, applicant_id):
    try:
        applicant = Application.objects.get(pk=applicant_id)
    except Application.DoesNotExist:
        raise Http404(f"{applicant_id} not found")

    pdf_data = generate_pdf(applicant)
    
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="지원서_{applicant.name}.pdf"'

    return response

# 여러 지원자 PDF ZIP 파일 다운로드
def download_pdf(request):
    applicant_ids = request.GET.getlist('applicants')

    if not applicant_ids:
        return HttpResponse("선택된 항목이 없습니다.", status=400)

    zip_buffer = io.BytesIO()

    # ZIP 파일 생성
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for applicant_id in applicant_ids:
            try:
                applicant = Application.objects.get(pk=applicant_id)
            except Application.DoesNotExist:
                raise Http404(f"Applicant with ID {applicant_id} not found")

            pdf_data = generate_pdf(applicant)
            zip_file.writestr(f"지원서_{applicant.name}.pdf", pdf_data)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="지원서.zip"'

    return response




def pass_document(request, applicant_id):
    if request.method == 'POST':
        try:
            applicant = Application.objects.get(pk=applicant_id)
            applicant.status = 'interview_scheduled'
            applicant.save()
            return JsonResponse({'success': True})
        except Application.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Applicant not Found'})

def fail_document(request, applicant_id):
    if request.method == 'POST':
        try:
            applicant = Application.objects.get(pk=applicant_id)
            applicant.status = 'submitted'
            applicant.save()
            return JsonResponse({'success': True})
        except Application.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Applicant not Found'})
    
def delete_recording(request, pk):
    applicant = get_object_or_404(Application, pk=pk)
    if request.method == 'POST':
        try:
            application = get_object_or_404(Application, pk=pk)
            recording = get_object_or_404(AudioRecording, application=applicant)
            recording.file.delete()  # 파일 시스템에서 파일 삭제
            recording.delete()  # DB에서 레코드 삭제
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)
    

def schedule(request):
    if request.user.is_authenticated:
        applicants = Application.objects.filter(~Q(status ='submitted'))
        schedules = Possible_date_list.objects.all()

        ctx = {"applicants": applicants, "schedules":schedules}
        return render(request, 'applicant/schedule.html', ctx)
    else:
        return redirect("accounts:login")
    

def auto_schedule(request):
    schedules = Possible_date_list.objects.all()

    # 스케줄 정보 리스트
    schedules_info = [[schedule.id, 0, schedule.max_possible_interview] for schedule in schedules]

    # 이미 배치된 인원 가능 인터뷰수에서 빼기
    scheduled_applicants = Application.objects.filter(interview_date__isnull=False)
    for scheduled_applicant in scheduled_applicants:
        for schedule_info in schedules_info:
            if schedule_info[0] == scheduled_applicant.interview_date.id:
                schedule_info[2] -= 1

    # 배치 가능 인터뷰 수 0 이하면 해당 스케줄 삭제
    for schedule_info in schedules_info:
        if schedule_info[2] <=0:
            schedules_info.remove(schedule_info)

    # 지원자 정보 리스트
    applicants = Application.objects.filter(interview_date__isnull=True)
    applicants_info = [[applicant.id, list(applicant.possible_date.all())] for applicant in applicants]

    # 스케줄 짜기 알고리즘
    while schedules_info:
        # 스케줄링 필요 인원 없으면 break
        if len(applicants) == 0:
            break

        # 삭제할 인원 인덱스 리스트
        del_applicant_index = []

        # 가능 시간대 0개 or 1개 인원 배치, 인기도 갱신
        for num in range(len(applicants_info)):
            # 0개
            if len(applicants_info[num][1]) == 0:
                del_applicant_index.append(num)
            # 1개
            elif len(applicants_info[num][1]) == 1:
                applicant = Application.objects.get(id=applicants_info[num][0])
                for schedule_info in schedules_info:
                    if schedule_info[0] == applicants_info[num][1][0].id:
                        if schedule_info[2] > 0:
                            applicant.interview_date = applicants_info[num][1][0]
                            applicant.save()
                            schedule_info[2] -= 1
                            break
                        else:
                            break

                del_applicant_index.append(num)
            # 인기도 갱신
            else:
                for possible_date in applicants_info[num][1]:
                    for schedule_info in schedules_info:
                        if schedule_info[0] == possible_date.id:
                            schedule_info[1] += 1
    
        # 배치 안되거나 배치한 애들 목록에서 삭제
        for index in sorted(del_applicant_index, reverse=True):
            del applicants_info[index]
        
        del_applicant_index = []    
        # 인기도 - 남는자리수 작은 시간대 추출
        popularity = 1000000
        low_pop_schedule = []
        for schedule_info in schedules_info:
            my_popularity = schedule_info[1] - schedule_info[2]
            if popularity > my_popularity:
                popularity = my_popularity
                low_pop_schedule = schedule_info

        # 인기도 - 남는자리수 낮은 시간대 배치
        for num in range(len(applicants_info)):
            for possible_date in applicants_info[num][1]:
                if possible_date.id == low_pop_schedule[0]:
                    if low_pop_schedule[2] > 0:
                        applicant = Application.objects.get(id=applicants_info[num][0])
                        applicant.interview_date = possible_date
                        applicant.save()
                        del_applicant_index.append(num)
                        low_pop_schedule[2] -=1
                    applicants_info[num][1].remove(possible_date)

        # 배치된 인원 및 시간대 삭제
        for index in sorted(del_applicant_index, reverse=True):
            del applicants_info[index]
        del_applicant_index = []

        schedules_info.remove(low_pop_schedule)
                    
    return redirect('applicants:schedule')
    

def schedule_update(request, pk):
    applicant = get_object_or_404(Application, id=pk)

    if request.method == 'POST':
        date_id = request.POST.get('selectDate')
        time_value = request.POST.get('selectTime')

        possible_date = get_object_or_404(Possible_date_list, id=date_id)
        
        # Parsing the time value from string to time object
        print(time_value[0])
        interview_time = f'{time_value[0]}{time_value[1]}:{time_value[3]}{time_value[4]}'
        if time_value[0] == '-':
            applicant.interview_time = None
        else:
            applicant.interview_time = interview_time
        applicant.interview_date = possible_date
        applicant.save()

        return redirect('applicants:schedule')

    return redirect('applicants:schedule')

def profile(request, pk):
    if request.user.is_authenticated:
        applicant = get_object_or_404(Application, pk=pk)
        answers = Answer.objects.filter(application=applicant)
        # 리코딩
        recording = AudioRecording.objects.filter(application=applicant).first()

        # 코멘트
        comments = Comment.objects.filter(application=applicant).order_by('created_at')
        form = CommentForm()

        # 질문 생성
        questions = individualQuestion.objects.filter(application=applicant).order_by('created_at')
        question_form = QuestionForm()
        answer_form = AnswerForm()

        # 공통 질문 템플릿 및 질문 가져오기
        common_template = InterviewTemplate.objects.get(is_default=True)
        common_questions = InterviewQuestion.objects.filter(template=common_template)
        
        # 녹음 파일 업로드 처리
        if request.method == 'POST' and request.FILES.get('audio_data'):
            try:
                # AudioRecording 객체 생성 또는 가져오기
                recording, created = AudioRecording.objects.get_or_create(application=applicant)
                recording.file = request.FILES['audio_data']
                recording.save()
                return JsonResponse({'file_url': recording.file.url})
            except Exception as e:
                print(f"Error saving file: {e}")
                return JsonResponse({'error': f'File upload failed: {str(e)}'}, status=500)

        ctx = {
            'applicant': applicant,
            'answers': answers,
            'recording_exists': recording is not None,
            'comments': comments,
            'form': form,
            'questions': questions,
            'question_form': question_form,
            'answer_form': answer_form,
            'common_questions': common_questions,
        }
        
        return render(request, 'applicant/profile.html', ctx)
    else:
        return redirect("accounts:login")
    

def comment(request, pk):
    applicant = get_object_or_404(Application, pk=pk)
    answers = Answer.objects.filter(application=applicant)
    comments = Comment.objects.filter(application=applicant).order_by('created_at')
    form = CommentForm()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        if request.method == "GET":
            ctx = {
            'applicant_id':pk,
            'applicant': applicant,
            'answers': answers,
            'comments': comments,
            'form': form,
            }
            return render(request, 'applicant/comments.html', ctx)

        # POST 일때
        else:
            interviewer = get_object_or_404(Interviewer, email=request.user.email)  # 인터뷰어 객체 가져오기
            form = CommentForm(request.POST)
            if form.is_valid():
                comment = form.save(commit=False)
                comment.application = applicant
                comment.interviewer = interviewer
                comment.save()
                return JsonResponse({
                    'success': True,
                    'comment': {
                        'text': comment.text,
                        'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                        'interviewer': interviewer.name,  # 인터뷰어 이메일 반환
                        'id': comment.id
                    }
                })
    return JsonResponse({'success': False, 'error': 'Invalid form submission', 'form_errors': form.errors.as_json()})
    

def delete_comment(request, pk, comment_id):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            comment = Comment.objects.get(pk=comment_id, application_id=pk)
            comment.delete()
            return JsonResponse({'success': True})
        except individualQuestion.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Question does not exist.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

def question(request, pk):
    applicant = get_object_or_404(Application, pk=pk)
    answers = Answer.objects.filter(application=applicant)
    questions = individualQuestion.objects.filter(application=applicant).order_by('created_at')
    question_form = QuestionForm()
    answer_form = AnswerForm()

    # 공통 질문 템플릿 및 질문 가져오기
    common_template = InterviewTemplate.objects.get(is_default=True)
    common_questions = InterviewQuestion.objects.filter(template=common_template)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        if request.method == "GET":
            ctx = {
                'applicant': applicant,
                'answers': answers,
                'questions': questions,
                'question_form': question_form,
                'answer_form': answer_form,
                'common_questions': common_questions,
            }
            return render(request, 'applicant/questions.html', ctx)    
        
        else:
            if 'question_submit' in request.POST:
                interviewer = get_object_or_404(Interviewer, email=request.user.email)
                question_form = QuestionForm(request.POST)
                if question_form.is_valid():
                    question = question_form.save(commit=False)
                    question.application = applicant
                    question.interviewer = interviewer
                    question.save()
                    return JsonResponse({
                        'success': True,
                        'question': {
                            'text': question.text,
                            'created_at': question.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                            'interviewer': interviewer.name,
                            'id': question.id
                        }
                    })
                else:
                    # 폼이 유효하지 않은 경우 오류 메시지 반환
                    return JsonResponse({'success': False, 'error': 'Invalid form submission', 'form_errors': question_form.errors.as_json()})
            
            elif 'answer_submit' in request.POST:
                interviewer = get_object_or_404(Interviewer, email=request.user.email)
                answer_form = AnswerForm(request.POST)
                if answer_form.is_valid():
                    answer = answer_form.save(commit=False)
                    answer.application = applicant
                    answer.interviewer = interviewer
                    answer.question_id = request.POST.get('question_id')
                    answer.save()
                    return JsonResponse({
                        'success': True,
                        'answer': {
                            'text': answer.text,
                            'created_at': answer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                            'interviewer': interviewer.email,
                            'id': answer.id
                        }
                    })
                else:
                    # 폼이 유효하지 않은 경우 오류 메시지 반환
                    return JsonResponse({'success': False, 'error': 'Invalid form submission', 'form_errors': question_form.errors.as_json()})
            
            else:
                print("iui")
                # 폼이 유효하지 않은 경우 오류 메시지 반환
                return JsonResponse({'success': False, 'error': 'Invalid form submission', 'form_errors': question_form.errors.as_json()})
            

def delete_question(request, pk, question_id):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            question = individualQuestion.objects.get(pk=question_id, application_id=pk)
            question.delete()
            return JsonResponse({'success': True})
        except individualQuestion.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Question does not exist.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

def delete_answer(request, pk, answer_id):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            answer = individualAnswer.objects.get(pk=answer_id, application_id=pk)
            answer.delete()
            return JsonResponse({'success': True})
        except individualQuestion.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Question does not exist.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

def applicant_rankings(req):
    applications = Application.objects.annotate(
        total_score=Coalesce(Sum('evaluations__total_score', filter=models.Q(evaluations__is_submitted=True)),0) # Evaluation모델을 역참조
    ).order_by('-total_score')
    interview_teams = InterviewTeam.objects.all()
    for interview_team in interview_teams:
        score_list = []
        for application in applications:
            # 팀을 자동설정 해주는 로직
            if set(interview_team.members.all()) == set(application.interviewer.all()):
                application.interview_team = interview_team
                application.save()
            # 설정된 팀을 가지고 
            if interview_team == application.interview_team:
                score_list.append(application.total_score)
                
        if len(score_list) != 0:
            interview_team.average_score = round(sum(score_list)/len(score_list),2)
            interview_team.save()
    context = {
        'applications': applications,
        'interview_teams': interview_teams,
    }
    return render(req, 'applicant_rankings.html', context)

## 지원 ##
@login_required(login_url='/applicants/login/')
def apply(request, pk):
    template = ApplicationTemplate.objects.get(id=pk)

    final_application = Application.objects.filter(template=template, applicant=request.user, is_drafted=False).first()
    if final_application:
        # 최종 제출된 지원서가 있으면 지원 완료 페이지로 리디렉션
        return redirect('applicants:apply_result')

    # GET 요청일 때 임시 저장된 지원서를 확인하여 `load_draft`로 리디렉션
    if request.method == 'GET':
        try:
            draft_application = Application.objects.get(template=template, applicant=request.user, is_drafted=True)
            return redirect('applicants:load_draft', pk=pk)
        except Application.DoesNotExist:
            pass  # 임시 저장된 지원서가 없으면 새로운 작성 페이지로 계속 진행

        # 새로운 작성 페이지를 위한 폼 생성
        form = ApplyForm()
        context = {
            'form': form,
            'template': template,
        }
        return render(request, 'for_applicant/write_apply.html', context)

    # POST 요청일 때 최종 제출을 처리
    if request.method == 'POST':
        form = ApplyForm(request.POST, request.FILES)
        if form.is_valid():
            applyContent = form.save(commit=False)
            applyContent.template = template
            applyContent.is_drafted = False
            applyContent.applicant = request.user
            applyContent.save()
            form.save_m2m()

            answers = {}
            for question in template.questions.all():
                answer_text = request.POST.get(f'answer_{question.id}')
                uploaded_file = request.FILES.get(f'file_{question.id}')

                answer, created = Answer.objects.get_or_create(
                    application=applyContent,
                    question=question
                )

                # 답변을 업데이트
                if answer_text:
                    answer.answer_text = answer_text
                    answers[question.id] = answer_text  # 기존 로직의 answers에 반영
                if uploaded_file:
                    answer.file_upload = uploaded_file

                # 최종 제출이므로 임시 저장 상태를 해제
                answer.is_drafted = False
                answer.save()
                
            transaction.on_commit(lambda: process_application.apply_async(args=(applyContent.id, answers), countdown=5))

            name = form.cleaned_data['name']
            phone_number = form.cleaned_data['phone_number']
            request.session['name'] = name
            request.session['phone_number'] = phone_number
            request.session['submitted'] = True

            return redirect('applicants:apply_result')
        else:
            print("폼 검증 실패:", form.errors)

        # 폼이 유효하지 않은 경우, 폼과 함께 다시 렌더링
        context = {
            'form': form,
            'template': template,
        }
        return render(request, 'for_applicant/write_apply.html', context)


#임시 저장 
@login_required
def save_draft(request, pk):
    template = ApplicationTemplate.objects.get(id=pk)

    try:
        applyContent = Application.objects.get(template=template, applicant=request.user, is_drafted=True)
    except Application.DoesNotExist:
        applyContent = None   # 없으면 빈 폼으로 시작

    if request.method == 'POST':
        form = ApplyForm(request.POST, request.FILES, instance=applyContent)
        if form.is_valid():
            applyContent = form.save(commit=False)
            applyContent.template = template
            applyContent.is_drafted = True  # 임시 저장 상태 설정
            applyContent.applicant = request.user
            applyContent.save()
            form.save_m2m()


            # 새 답변 저장
            for question in template.questions.all():
                answer_text = request.POST.get(f'answer_{question.id}')
                uploaded_file = request.FILES.get(f'file_{question.id}')

                # 기존 답변이 있는지 확인하고 있으면 업데이트, 없으면 새로 저장
                answer, created = Answer.objects.get_or_create(
                    application=applyContent,
                    question=question,
                    defaults={'answer_text': answer_text, 'file_upload': uploaded_file}
                )
                if not created:
                    # 기존 답변이 있다면 업데이트
                    answer.answer_text = answer_text
                    answer.file_upload = uploaded_file
                    answer.save()
                
                print(f"답변 저장됨: 질문 ID={question.id}, 답변 텍스트={answer.answer_text}, 파일={answer.file_upload}")

            answers = {str(answer.question.id): answer.answer_text for answer in applyContent.answers.all()}
            print(f"생성된 딕셔너리 answers: {answers}")

            return JsonResponse({'status': 'draft_saved', 'application_id': applyContent.id})
        else:
            return JsonResponse({'status': 'error', 'message': 'Form is invalid'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@login_required
def load_draft(request, pk):
    if request.method == 'GET':
        template = ApplicationTemplate.objects.get(id=pk)

        try:
            applyContent = Application.objects.get(template=template, applicant=request.user, is_drafted=True)
            form = ApplyForm(instance=applyContent)
            answers = {str(answer.question.id): answer.answer_text for answer in applyContent.answers.all()}
            
        except Application.DoesNotExist:
            print("Application 객체를 찾을 수 없습니다.")
            form = ApplyForm()
            answers = {}

        context = {
            'form': form,
            'template': template,
            'answers': answers,
        }
        return render(request, 'for_applicant/write_apply.html', context)

    # 만약 POST 등의 다른 요청이 오면 허용되지 않은 메서드로 처리
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

def apply_timeover(request):
    return render(request, 'for_applicant/timeover.html')

def download_default_excel(request):
    default_evaluate = EvaluationTemplate.objects.filter(is_default=True).first()
    if not default_evaluate:
        return HttpResponse("기본 템플릿이 설정되지 않았습니다.", status=404)

    applications = Application.objects.filter(evaluation_template=default_evaluate).select_related('interview_team').prefetch_related('interviewer', 'comments', 'evaluations')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "평가표"

    # 제목, 설명
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{default_evaluate.title}"
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{default_evaluate.description}"

    # 테이블 헤더 설정
    headers = ["면접팀", "면접자", "지원자", "학교", "전공"]
    question_headers = [f"{idx + 1}번" for idx, question in enumerate(default_evaluate.questions.all())]
    headers.extend(question_headers)
    headers.append("총 점수")
    headers.append("코멘트")
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}3'] = header

    row_num = 4  
    for application in applications:
        interview_team = application.interview_team
        applicant_name = application.name
        school = application.school if application.school else ""
        major = application.major if application.major else ""
        
        # 면접 팀의 모든 멤버 가져오기
        interviewers = interview_team.members.all() if interview_team else []

        # 지원자 정보를 첫 번째 행에만 기록
        first_row = True
        for interviewer in interviewers:
            # 해당 면접관의 평가가 있을 경우 불러오기
            evaluation = application.evaluations.filter(interviewer=interviewer).first()
            ws[f'A{row_num}'] = interview_team.team_name if interview_team and first_row else ""  # 면접팀 
            ws[f'B{row_num}'] = interviewer.name  # 면접관 
            ws[f'C{row_num}'] = applicant_name if first_row else ""  # 지원자
            ws[f'D{row_num}'] = school if first_row else ""  # 학교
            ws[f'E{row_num}'] = major if first_row else ""  # 전공
            first_row = False  

            # 질문 개별 점수
            for idx, question in enumerate(default_evaluate.questions.all(), start=1):
                col_letter = get_column_letter(5 + idx)
                score = evaluation.scores.filter(question=question).first().score if evaluation and evaluation.scores.filter(question=question).exists() else 0
                ws[f'{col_letter}{row_num}'] = score

            # 질문 총 점수
            total_score_col = get_column_letter(5 + len(question_headers) + 1)
            ws[f'{total_score_col}{row_num}'] = evaluation.total_score if evaluation else 0
    
            # 코멘트
            comments_col = get_column_letter(5 + len(question_headers) + 2)
            comment = application.comments.filter(interviewer=interviewer).first()
            ws[f'{comments_col}{row_num}'] = comment.text if comment else ""
            row_num += 1  

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={default_evaluate.title}_평가표.xlsx'
    
    wb.save(response)
    return response

def apply_result(request):
    # 현재 로그인된 사용자 정보와 관련된 제출 상태를 DB에서 조회
    try:
        application = Application.objects.get(applicant=request.user, is_drafted=False)
        submitted = True
    except Application.DoesNotExist:
        submitted = False
    context = {
        'submitted': submitted,
    }
    return render(request, 'for_applicant/apply_result.html', context) if submitted else redirect('applicants:initial')